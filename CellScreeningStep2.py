import os
from io import BytesIO
from typing import Dict, List, Optional

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from sklearn.preprocessing import StandardScaler

from analysis_config import STEP2_COLUMNS, MIN_CLUSTER_SIZE, DEFAULT_WEIGHTS


DEFAULT_OUTPUT_DIR = "Results"
DEFAULT_CLUSTER_INDEX = 1
ANALYSIS_COLUMNS = STEP2_COLUMNS

def _calculate_ranges(df: pd.DataFrame, cols: List[str]) -> Dict[str, Dict[str, float]]:
    ranges = {}
    for col in cols:
        q1 = df[col].quantile(0.33)
        q2 = df[col].quantile(0.66)
        vmin, vmax = df[col].min(), df[col].max()
        if q1 == q2:
            q1 = vmin + (vmax - vmin) / 3
            q2 = vmin + (vmax - vmin) * 2 / 3
        ranges[col] = {"low": vmin, "mid1": q1, "mid2": q2, "high": vmax}
    return ranges


def _membership_percentile(x, col, ranges, df_col):
    low, mid1, mid2, high = (
        ranges[col]["low"],
        ranges[col]["mid1"],
        ranges[col]["mid2"],
        ranges[col]["high"],
    )

    if low <= x < mid1:
        low_pct = (df_col[(df_col >= low) & (df_col < mid1)] <= x).mean() * 100
    else:
        low_pct = 0.0

    if mid1 <= x <= mid2:
        med_pct = (df_col[(df_col >= mid1) & (df_col <= mid2)] <= x).mean() * 100
    else:
        med_pct = 0.0

    if mid2 < x <= high:
        high_pct = (df_col[(df_col > mid2) & (df_col <= high)] <= x).mean() * 100
    else:
        high_pct = 0.0

    return low_pct, med_pct, high_pct


def _select_top_pack_num(df_fis: pd.DataFrame) -> pd.DataFrame:
    priority_order = ["Low", "Med", "High"]
    priority_map = {cls: i for i, cls in enumerate(priority_order)}

    selected_list = []
    remaining = MIN_CLUSTER_SIZE

    for cls in priority_order:
        if remaining <= 0:
            break
        cls_df = df_fis[df_fis["FIS_Class"] == cls].copy()
        pct_col = f"{cls}_mean(%)"
        if pct_col not in cls_df.columns or cls_df.empty:
            continue
        cls_df = cls_df.sort_values(by=pct_col, ascending=True)
        take_n = min(remaining, len(cls_df))
        if take_n > 0:
            selected_list.append(cls_df.head(take_n))
            remaining -= take_n

    if remaining > 0:
        already_selected = (
            pd.concat(selected_list, ignore_index=True)["Lot Number"].tolist()
            if selected_list
            else []
        )
        leftover = df_fis[~df_fis["Lot Number"].isin(already_selected)].copy()
        if not leftover.empty:
            def get_class_pct_val(row):
                cls = row["FIS_Class"]
                pct_col = f"{cls}_mean(%)"
                val = row.get(pct_col, np.nan)
                return val if pd.notna(val) else 1e9

            leftover["_class_priority"] = leftover["FIS_Class"].map(priority_map)
            leftover["_class_pct"] = leftover.apply(get_class_pct_val, axis=1)
            leftover = leftover.sort_values(by=["_class_priority", "_class_pct"], ascending=[True, True])
            to_take = leftover.head(remaining)
            selected_list.append(to_take)

    if selected_list:
        return pd.concat(selected_list, ignore_index=True).reset_index(drop=True)
    return pd.DataFrame(columns=df_fis.columns)


def run_step2(
    cs1_file: str,
    cluster_index: Optional[int] = DEFAULT_CLUSTER_INDEX,
    output_dir: str = DEFAULT_OUTPUT_DIR,
    cols: Optional[List[str]] = None,
    worst_cluster: Optional[int] = None,
) -> str:
    #cols = cols or ACIR_COLUMNS
    cols = cols or ANALYSIS_COLUMNS
    cluster_index = cluster_index if cluster_index is not None else DEFAULT_CLUSTER_INDEX
    sheet_name = f"Cluster{cluster_index}"

    print(f"[Step2] Loading '{cs1_file}' sheet '{sheet_name}'...")
    df = pd.read_excel(cs1_file, sheet_name=sheet_name)
    available_cols = [c for c in cols if c in df.columns]
    missing_cols = [c for c in cols if c not in df.columns]
    if missing_cols:
        print(f"[Step2][WARN] Missing columns skipped: {missing_cols}")
    if not available_cols:
        # fallback: pick first 2 numeric-like columns (excluding Lot Number)
        candidate_cols = [c for c in df.columns if c != "Lot Number"]
        numeric_cols = []
        for c in candidate_cols:
            coerced = pd.to_numeric(df[c], errors="coerce")
            if coerced.notna().any():
                numeric_cols.append(c)
        if not numeric_cols:
            raise ValueError(f"[Step2] No analysis columns found. Requested: {cols}")
        available_cols = numeric_cols[:2]
        print(f"[Step2][WARN] Using fallback columns: {available_cols}")
    df[available_cols] = df[available_cols].apply(pd.to_numeric, errors="coerce")
    cols = available_cols
    print(f"[Step2] Loaded {len(df)} rows for analysis columns: {cols}")

    worst_cells_df = None
    if worst_cluster is not None:
        worst_sheet = f"Cluster{worst_cluster}"
        print(f"[Step2] Loading worst cluster data from sheet '{worst_sheet}'...")
        try:
            df_worst = pd.read_excel(cs1_file, sheet_name=worst_sheet)
            if "Lot Number" in df_worst.columns:
                # Include full feature columns for worst cluster, sorted by lot number for readability
                worst_cells_df = df_worst.sort_values(by="Lot Number").reset_index(drop=True)
            else:
                print(f"[Step2][WARN] '{worst_sheet}' does not contain 'Lot Number' column. Skipping Worst Cells sheet.")
        except Exception as exc:
            print(f"[Step2][WARN] Failed to load sheet {worst_sheet} for Worst Cells: {exc}")

    print("[Step2] Calculating percentile ranges and membership scores...")
    ranges = _calculate_ranges(df, cols)
    df_fis = df[["Lot Number"] + cols].copy()
    for col in cols:
        df_fis[[f"{col}_Low_pct(%)", f"{col}_Med_pct(%)", f"{col}_High_pct(%)"]] = df_fis[col].apply(
            lambda x: pd.Series(_membership_percentile(x, col, ranges, df[col]))
        )

    df_fis["Low_mean(%)"] = df_fis[[f"{c}_Low_pct(%)" for c in cols]].mean(axis=1)
    df_fis["Med_mean(%)"] = df_fis[[f"{c}_Med_pct(%)" for c in cols]].mean(axis=1)
    df_fis["High_mean(%)"] = df_fis[[f"{c}_High_pct(%)" for c in cols]].mean(axis=1)

    def pick_fis_class(row):
        vals = [row["Low_mean(%)"], row["Med_mean(%)"], row["High_mean(%)"]]
        idx = int(np.argmax(vals))
        return ["Low", "Med", "High"][idx]

    df_fis["FIS_Class"] = df_fis.apply(pick_fis_class, axis=1)

    df_ranges = pd.DataFrame(columns=["Item", "Low_range", "Med_range", "High_range"])
    for col in cols:
        low_range = f"{ranges[col]['low']:.2f} ~ {ranges[col]['mid1'] - 0.01:.2f}"
        med_range = f"{ranges[col]['mid1']:.2f} ~ {ranges[col]['mid2']:.2f}"
        high_range = f"{ranges[col]['mid2'] + 0.01:.2f} ~ {ranges[col]['high']:.2f}"
        df_ranges = pd.concat(
            [
                df_ranges,
                pd.DataFrame(
                    {
                        "Item": [col],
                        "Low_range": [low_range],
                        "Med_range": [med_range],
                        "High_range": [high_range],
                    }
                ),
            ],
            ignore_index=True,
        )

    pack_num = MIN_CLUSTER_SIZE
    print(f"[Step2] Selecting top {pack_num} cells based on FIS results...")
    df_selected_pack_num = _select_top_pack_num(df_fis.copy())
    selected_lot_nums = df_selected_pack_num["Lot Number"].unique().tolist()
    df_selected_raw = df[df["Lot Number"].isin(selected_lot_nums)].copy()
    print(f"[Step2] Selected {len(df_selected_pack_num)} cells for Best_{pack_num} sheets.")

    best_suffix = f"({cluster_index})" if cluster_index is not None else ""
    worst_suffix = f"({worst_cluster})" if worst_cluster is not None else ""
    best_sheet_name = (
        f"Best_{MIN_CLUSTER_SIZE}cells{best_suffix}" if best_suffix else f"Best_{MIN_CLUSTER_SIZE}cells"
    )
    best_raw_sheet_name = (
        f"Best_{MIN_CLUSTER_SIZE}cells_raw{best_suffix}"
        if best_suffix
        else f"Best_{MIN_CLUSTER_SIZE}cells_raw"
    )
    worst_sheet_name = f"Worst Cells{worst_suffix}" if worst_suffix else "Worst Cells"

    output_path = os.path.join(output_dir, "Step2_Results.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Cluster_Data", index=False)
        df_ranges.to_excel(writer, sheet_name="Ranges", index=False)
        df_fis.to_excel(writer, sheet_name="FIS_Result", index=False)
        df_selected_pack_num.to_excel(writer, sheet_name=best_sheet_name, index=False)
        df_selected_raw.to_excel(writer, sheet_name=best_raw_sheet_name, index=False)
        if worst_cells_df is not None:
            worst_cells_df.to_excel(writer, sheet_name=worst_sheet_name, index=False)

        fis_class_counts = df_fis["FIS_Class"].value_counts().reindex(["Low", "Med", "High"], fill_value=0)
        fig_class, ax_class = plt.subplots(figsize=(5, 4))
        ax_class.pie(
            fis_class_counts,
            labels=fis_class_counts.index,
            autopct="%1.1f%%",
            colors=["#4daf4a", "#ff7f00", "#377eb8"],
            startangle=90,
        )
        ax_class.set_title("FIS Class Distribution")
        class_img = BytesIO()
        plt.savefig(class_img, format="png", dpi=150)
        plt.close(fig_class)
        class_img.seek(0)
        class_sheet = writer.book.create_sheet("FIS_Class_Distribution")
        class_image = XLImage(class_img)
        class_image.width, class_image.height = 480, 360
        class_sheet.add_image(class_image, "A1")

        fig_ranges, axes_ranges = plt.subplots(len(cols), 1, figsize=(6, 3 * len(cols)))
        axes_ranges = np.array(axes_ranges).reshape(-1)
        for ax, col in zip(axes_ranges, cols):
            ax.plot(df[col].values, label=col, color="#1f77b4")
            ax.axhspan(ranges[col]["low"], ranges[col]["mid1"], color="#4daf4a", alpha=0.1, label="Low Range")
            ax.axhspan(ranges[col]["mid1"], ranges[col]["mid2"], color="#ff7f00", alpha=0.1, label="Mid Range")
            ax.axhspan(ranges[col]["mid2"], ranges[col]["high"], color="#377eb8", alpha=0.1, label="High Range")
            ax.set_title(f"{col} Distribution with Ranges")
            ax.set_ylabel(col)
            ax.legend(loc="best", fontsize=8)
        plt.tight_layout()
        range_img = BytesIO()
        plt.savefig(range_img, format="png", dpi=150)
        plt.close(fig_ranges)
        range_img.seek(0)
        ranges_sheet = writer.book.create_sheet("Ranges_Chart")
        range_image = XLImage(range_img)
        range_image.width, range_image.height = 640, 320 * len(cols)
        ranges_sheet.add_image(range_image, "A1")

        if len(cols) >= 2:
            class_color_map = {"Low": "#4daf4a", "Med": "#ff7f00", "High": "#377eb8"}
            colors = df_fis["FIS_Class"].map(class_color_map)

            # 1) PCA 2D scatter
            fig_pca, ax_pca = plt.subplots(figsize=(6, 5))
            scaler = StandardScaler()
            X = scaler.fit_transform(df[cols].apply(pd.to_numeric, errors="coerce").fillna(0))
            # simple 2D PCA via SVD
            X_centered = X - X.mean(axis=0, keepdims=True)
            u, s, vt = np.linalg.svd(X_centered, full_matrices=False)
            pc2 = u[:, :2] * s[:2]
            cluster_label = (
                f"Cluster {cluster_index}" if cluster_index is not None else "Cluster"
            )
            ax_pca.scatter(pc2[:, 0], pc2[:, 1], c=colors, alpha=0.6, label=cluster_label)
            if not df_selected_raw.empty:
                idx_selected = df["Lot Number"].isin(df_selected_raw["Lot Number"])
                ax_pca.scatter(
                    pc2[idx_selected, 0],
                    pc2[idx_selected, 1],
                    color="red",
                    label=f"Best {MIN_CLUSTER_SIZE}",
                    edgecolors="white",
                    s=15,
                )
            ax_pca.set_xlabel("PC1")
            ax_pca.set_ylabel("PC2")
            class_counts = df_fis["FIS_Class"].value_counts().reindex(["Low", "Med", "High"], fill_value=0)
            ax_pca.set_title(f"PCA 2D (Best {MIN_CLUSTER_SIZE} Highlight)")
            from matplotlib.lines import Line2D
            base_handles, base_labels = ax_pca.get_legend_handles_labels()
            class_handles = [
                Line2D(
                    [0],
                    [0],
                    marker="o",
                    color="w",
                    label=f"Low ({class_counts['Low']})",
                    markerfacecolor="#4daf4a",
                    markersize=7,
                ),
                Line2D(
                    [0],
                    [0],
                    marker="o",
                    color="w",
                    label=f"Med ({class_counts['Med']})",
                    markerfacecolor="#ff7f00",
                    markersize=7,
                ),
                Line2D(
                    [0],
                    [0],
                    marker="o",
                    color="w",
                    label=f"High ({class_counts['High']})",
                    markerfacecolor="#377eb8",
                    markersize=7,
                ),
            ]
            ax_pca.legend(
                base_handles + class_handles,
                base_labels + [h.get_label() for h in class_handles],
                loc="best",
            )
            plt.tight_layout()
            pca_img = BytesIO()
            plt.savefig(pca_img, format="png", dpi=150)
            plt.close(fig_pca)
            pca_img.seek(0)
            pca_sheet = writer.book.create_sheet(f"Best{MIN_CLUSTER_SIZE}_PCA2D")
            pca_image = XLImage(pca_img)
            pca_image.width, pca_image.height = 640, 480
            pca_sheet.add_image(pca_image, "A1")

            # 2) Top-2 weight columns scatter
            weight_sorted = sorted(DEFAULT_WEIGHTS.items(), key=lambda x: x[1], reverse=True)
            top2 = [col for col, _ in weight_sorted if col in cols][:2]
            if len(top2) == 2:
                fig_scatter, ax_scatter = plt.subplots(figsize=(6, 5))
                ax_scatter.scatter(
                    df[top2[0]],
                    df[top2[1]],
                    c=colors,
                    alpha=0.6,
                    label=cluster_label,
                )
                if not df_selected_raw.empty:
                    ax_scatter.scatter(
                        df_selected_raw[top2[0]],
                        df_selected_raw[top2[1]],
                        color="red",
                        label=f"Best {MIN_CLUSTER_SIZE}",
                        edgecolors="white",
                        s=15,
                    )
                ax_scatter.set_xlabel(top2[0])
                ax_scatter.set_ylabel(top2[1])
                ax_scatter.set_title(f"Top-2 Weighted Scatter (Best {MIN_CLUSTER_SIZE})")
                base_handles, base_labels = ax_scatter.get_legend_handles_labels()
                class_handles = [
                    Line2D(
                        [0],
                        [0],
                        marker="o",
                        color="w",
                        label=f"Low ({class_counts['Low']})",
                        markerfacecolor="#4daf4a",
                        markersize=7,
                    ),
                    Line2D(
                        [0],
                        [0],
                        marker="o",
                        color="w",
                        label=f"Med ({class_counts['Med']})",
                        markerfacecolor="#ff7f00",
                        markersize=7,
                    ),
                    Line2D(
                        [0],
                        [0],
                        marker="o",
                        color="w",
                        label=f"High ({class_counts['High']})",
                        markerfacecolor="#377eb8",
                        markersize=7,
                    ),
                ]
                ax_scatter.legend(
                    base_handles + class_handles,
                    base_labels + [h.get_label() for h in class_handles],
                    loc="best",
                )
                plt.tight_layout()
                scatter_img = BytesIO()
                plt.savefig(scatter_img, format="png", dpi=150)
                plt.close(fig_scatter)
                scatter_img.seek(0)
                scatter_sheet = writer.book.create_sheet(f"Best{MIN_CLUSTER_SIZE}_Top2")
                scatter_image = XLImage(scatter_img)
                scatter_image.width, scatter_image.height = 640, 480
                scatter_sheet.add_image(scatter_image, "A1")
            else:
                print("[Step2][WARN] Top-2 weighted columns not found for scatter plot.")

    print(f"[Step2] 저장 완료: {output_path}")
    return output_path


__all__ = ["run_step2"]
