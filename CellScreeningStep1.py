import os
from io import BytesIO
from typing import Dict, List

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from sklearn.metrics import adjusted_rand_score, calinski_harabasz_score, davies_bouldin_score
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from k_means_constrained import KMeansConstrained


COLUMNS_TO_ANALYZE = {
    "All item": [
        "Capacity(Ah)",
        "Weight(g)",
        "Height(mm)",
        "Width(mm)",
        "Initial Voltage(V)",
        "100% Voltage(V)",
        "0% Voltage(V)",
        "50% Voltage(V)",
        "Initial ACIR(mΩ)",
        "100% ACIR(mΩ)",
        "0% ACIR(mΩ)",
        "50% ACIR(mΩ)",
    ]
}

STD_COLS = [
    "Capacity(Ah)",
    "Weight(g)",
    "Height(mm)",
    "Width(mm)",
    "Initial Voltage(V)",
    "100% Voltage(V)",
    "0% Voltage(V)",
    "50% Voltage(V)",
    "Initial ACIR(mΩ)",
    "100% ACIR(mΩ)",
    "0% ACIR(mΩ)",
    "50% ACIR(mΩ)",
]

DEFAULT_OUTPUT_DIR = "Results"
MIN_CLUSTER_SIZE = 72

# 권장 가중치 설정 (0 ~ 5)
DEFAULT_WEIGHTS = {
    # Tier 1: 성능/안전 직결
    "Capacity(Ah)": 3.0,
    "Initial Voltage(V)": 3.0,
    "Initial ACIR(mΩ)": 2.5,
    
    # Tier 2: 운영 특성
    "100% ACIR(mΩ)": 2.0,
    "0% ACIR(mΩ)": 2.0,
    "50% ACIR(mΩ)": 2.0,
    
    "100% Voltage(V)": 1.0,
    "0% Voltage(V)": 1.0,
    "50% Voltage(V)": 1.0,
    
    # Tier 3: 부가 특성
    "Weight(g)": 1.5,
    
    # Tier 4: 물리적 호환성
    "Height(mm)": 0.5,
    "Width(mm)": 0.5,
}

class BalancedKMeans:
    def __init__(self, n_clusters=10, max_iter=100, random_state=42, tol=1e-4):
        self.n_clusters = n_clusters
        self.max_iter = max_iter
        self.random_state = np.random.RandomState(random_state)
        self.tol = tol
        self.cluster_centers_ = None
        self.labels_ = None

    def _kmeans_plus_plus_init(self, X):
        n, d = X.shape
        centers = np.empty((self.n_clusters, d), dtype=float)
        idx = self.random_state.randint(0, n)
        centers[0] = X[idx]
        closest_dist_sq = np.sum((X - centers[0]) ** 2, axis=1)
        for c in range(1, self.n_clusters):
            probs = closest_dist_sq / closest_dist_sq.sum()
            idx = self.random_state.choice(n, p=probs)
            centers[c] = X[idx]
            new_dist_sq = np.sum((X - centers[c]) ** 2, axis=1)
            closest_dist_sq = np.minimum(closest_dist_sq, new_dist_sq)
        return centers

    def _balanced_assignment(self, X, centers, caps):
        n = X.shape[0]
        k = centers.shape[0]
        dists = np.linalg.norm(X[:, None, :] - centers[None, :, :], axis=2)
        order = np.argmin(dists, axis=1)
        base = dists[np.arange(n), order]
        idx_order = np.argsort(base)
        labels = -np.ones(n, dtype=int)
        caps_left = caps.copy()
        nearest_rank = np.argsort(dists, axis=1)
        for i in idx_order:
            for j in range(k):
                c = nearest_rank[i, j]
                if caps_left[c] > 0:
                    labels[i] = c
                    caps_left[c] -= 1
                    break
        if np.any(labels == -1):
            for i in np.where(labels == -1)[0]:
                c = np.argmax(caps_left)
                labels[i] = c
                caps_left[c] -= 1
        return labels

    def fit_predict(self, X):
        n = X.shape[0]
        k = self.n_clusters
        base = n // k
        r = n % k
        capacities = np.array([base + 1 if i < r else base for i in range(k)], dtype=int)
        centers = self._kmeans_plus_plus_init(X)
        labels_prev = None
        for _ in range(self.max_iter):
            labels = self._balanced_assignment(X, centers, capacities)
            new_centers = np.array([X[labels == c].mean(axis=0) for c in range(k)])
            shift = np.linalg.norm(new_centers - centers)
            centers = new_centers
            if labels_prev is not None and np.all(labels == labels_prev):
                break
            if shift < self.tol:
                break
            labels_prev = labels
        self.cluster_centers_ = centers
        self.labels_ = labels
        return labels


def _normalize_scores(scores: List[float], mode: str) -> np.ndarray:
    arr = np.array(scores)
    if arr.size == 0:
        return arr
    span = np.max(arr) - np.min(arr)
    if span == 0:
        return np.ones_like(arr)
    if mode == "dbi":
        return 1 - (arr - np.min(arr)) / span
    return (arr - np.min(arr)) / span


def compute_k_metrics(df: pd.DataFrame) -> Dict[str, Dict[str, object]]:
    results = {}
    for name, cols in COLUMNS_TO_ANALYZE.items():
        if any(col not in df.columns for col in cols):
            continue
        data = df[cols].dropna().values
        if data.size == 0:
            continue

        print(f"[Step1] Computing K metrics for '{name}' ({len(data)} samples)...")
        K_range = range(9, 15) ################## K ####################
        feasible_k = [k for k in K_range if k * MIN_CLUSTER_SIZE <= len(data)]
        if not feasible_k:
            raise ValueError(
                f"데이터 수({len(data)})로는 클러스터 최소 크기 {MIN_CLUSTER_SIZE}을 만족하는 k를 찾을 수 없습니다."
            )
        dbi_scores, chi_scores = [], []
        for k in feasible_k:
            print(f"[Step1]   - evaluating k={k}...")
            size_max = max(MIN_CLUSTER_SIZE, int(len(data) / k * 1.2))
            kmeans = KMeansConstrained(
                n_clusters=k,
                size_min=MIN_CLUSTER_SIZE,
                size_max=size_max,
                random_state=42,                #의미없는 시작값 0~100
            )
            labels = kmeans.fit_predict(data)
            dbi_scores.append(davies_bouldin_score(data, labels))       #DBI 점수 계산 (낮을수록 좋음 ↓)
            chi_scores.append(calinski_harabasz_score(data, labels))    #CHI 점수 계산 (높을수록 좋음 ↑)

        dbi_norm = _normalize_scores(dbi_scores, "dbi")
        chi_norm = _normalize_scores(chi_scores, "chi")
        combined_score = 0.5 * dbi_norm + 0.5 * chi_norm
        optimal_k_final = feasible_k[int(np.argmax(combined_score))]
        print(f"[Step1] [{name}] Optimal K (Combined): {optimal_k_final}")
        results[name] = {
            "k_values": list(feasible_k),
            "dbi_scores": dbi_scores,
            "chi_scores": chi_scores,
            "optimal_k_dbi": feasible_k[int(np.argmin(dbi_scores))],
            "optimal_k_chi": feasible_k[int(np.argmax(chi_scores))],
            "optimal_k_final": optimal_k_final,
        }
    return results


def _add_k_selection_sheet(workbook, k_results: Dict[str, Dict[str, object]]):
    ws = workbook.create_sheet("K_Selection_Result")
    ws.append(["분석 대상", "DBI k", "CHI k", "최종 추천 k"])

    for name, result in k_results.items():
        ws.append(
            [
                name,
                result["optimal_k_dbi"],
                result["optimal_k_chi"],
                result["optimal_k_final"],
            ]
        )
        fig, axs = plt.subplots(1, 2, figsize=(10, 4))
        axs[0].plot(result["k_values"], result["dbi_scores"], "o-", label="DBI (↓)")
        axs[0].axvline(result["optimal_k_dbi"], color="r", linestyle="--", label="Optimal DBI")
        axs[0].set_title("Davies-Bouldin Index")
        axs[0].set_xlabel("k")
        axs[0].legend()

        axs[1].plot(result["k_values"], result["chi_scores"], "o-", label="CHI (↑)")
        axs[1].axvline(result["optimal_k_chi"], color="r", linestyle="--", label="Optimal CHI")
        axs[1].set_title("Calinski-Harabasz Index")
        axs[1].set_xlabel("k")
        axs[1].legend()

        fig.suptitle(f"Optimal K Determination for {name}", fontsize=14)
        plt.tight_layout(rect=[0, 0, 1, 0.95])

        img_data = BytesIO()
        plt.savefig(img_data, format="png", bbox_inches="tight")
        plt.close(fig)
        img = XLImage(img_data)
        img.width, img.height = 600, 250
        ws.add_image(img, f"G{ws.max_row}")


def run_step1(
    cs0_file: str, 
    output_dir: str = DEFAULT_OUTPUT_DIR,
    weights: dict = None,  # 새 파라미터
    use_equal_weights: bool = False  # 균등 가중치 옵션
):
    #     
    df = pd.read_excel(cs0_file, sheet_name="Non_Outliers_List")
    k_results = compute_k_metrics(df)
    if "All item" not in k_results:
        raise ValueError("필요한 분석 항목 데이터를 찾을 수 없습니다.")

    k = k_results["All item"]["optimal_k_final"]
    id_col = "Lot Number"
    feature_cols = COLUMNS_TO_ANALYZE["All item"]

    df_use = df[[id_col] + feature_cols].dropna().reset_index(drop=True)

    # ========== 가중치 적용 부분 (클러스터링 포함) ==========
    if use_equal_weights:
        # 균등 가중치 (기존 방식)
        print("[Step1] 균등 가중치 사용 (모든 항목 1.0)")
        weights_to_use = {col: 1.0 for col in feature_cols}
    else:
        # 가중치 적용
        if weights is None:
            weights_to_use = DEFAULT_WEIGHTS
            print("[Step1] 가중치 적용:")
        else:
            weights_to_use = weights
            print("[Step1] 사용자 정의 가중치 적용:")

    # 가중치 출력
    for col in feature_cols:
        w = weights_to_use.get(col, 1.0)
        print(f"[Step1]    {col:25s}: {w:.1f}")
    # =====================================

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(df_use[feature_cols])
    weights_vec = np.array([weights_to_use.get(col, 1.0) for col in feature_cols])
    X_weighted = X_scaled * weights_vec
    pca = PCA(n_components=2, random_state=42)
    coords = pca.fit_transform(X_weighted) if X_weighted.shape[1] >= 2 else np.column_stack(
        [X_weighted[:, 0], np.zeros_like(X_weighted[:, 0])]
    )

    size_max = max(MIN_CLUSTER_SIZE, int(len(df_use) / k * 1.2))
    kmeans_weighted = KMeansConstrained(
        n_clusters=k,
        size_min=MIN_CLUSTER_SIZE,
        size_max=size_max,
        random_state=42,
    )
    labels = kmeans_weighted.fit_predict(X_weighted)

    kmeans_unweighted = KMeansConstrained(
        n_clusters=k,
        size_min=MIN_CLUSTER_SIZE,
        size_max=size_max,
        random_state=42,
    )
    labels_unweighted = kmeans_unweighted.fit_predict(X_scaled)
    df_use["cluster"] = labels
    df_out = df.merge(df_use[[id_col, "cluster"]], on=id_col, how="left")

    df_scaled_unweighted = df_use.copy()
    df_scaled_unweighted[feature_cols] = X_scaled
    df_scaled_unweighted["cluster"] = labels_unweighted

    df_scaled_weighted = df_use.copy()
    df_scaled_weighted[feature_cols] = X_weighted
    df_scaled_weighted["cluster"] = labels

    df_counts = df_use["cluster"].value_counts().sort_index().reset_index()
    df_counts.columns = ["cluster", "count"]

    # 표준편차 계산
    std_cols_use = [c for c in STD_COLS if c in df_out.columns]
    df_cluster_std = df_out.groupby("cluster")[std_cols_use].std().reset_index()
    
    # 순위 변환
    df_rank = df_cluster_std.copy()
    for col in std_cols_use:
        df_rank[col] = df_cluster_std[col].rank(method="min", ascending=True)

    scatter_plot_df = pd.DataFrame(
        {
            "Lot Number": df_use[id_col],
            "cluster": labels,
            "PC1": coords[:, 0],
            "PC2": coords[:, 1],
        }
    )

    # 가중치 적용 순위 합산
    df_rank["total_rank"] = sum(
        df_rank[col] * weights_to_use.get(col, 1.0)
        for col in std_cols_use
    )
    # =====================================
    
    # 최고 / 최악 클러스터 선정
    best_cluster = int(df_rank.loc[df_rank["total_rank"].idxmin(), "cluster"]) if len(df_rank) else None
    worst_cluster = int(df_rank.loc[df_rank["total_rank"].idxmax(), "cluster"]) if len(df_rank) else None
    print(f"[Step1] 가장 안정적인 클러스터: {best_cluster}")
    print(f"[Step1] 가장 변동성이 큰 클러스터: {worst_cluster}")
    
    # 가중치 정보 추가 저장
    df_weights = pd.DataFrame([
        {"항목": col, "가중치": weights_to_use.get(col, 1.0)}
        for col in std_cols_use
    ])

    compare_ari = adjusted_rand_score(labels_unweighted, labels)
    df_compare_summary = pd.DataFrame(
        [
            {"Metric": "Adjusted Rand Index", "Value": compare_ari},
            {"Metric": "Weighted Clusters (k)", "Value": k},
            {"Metric": "Unweighted Clusters (k)", "Value": k},
        ]
    )
    df_compare_matrix = pd.crosstab(
        labels_unweighted,
        labels,
        rownames=["Unweighted Cluster"],
        colnames=["Weighted Cluster"],
    )
    
    os.makedirs(output_dir, exist_ok=True)
    base_path = os.path.join(output_dir, "Step1_Results.xlsx")
    cs1_path = base_path
    if os.path.exists(base_path):
        try:
            os.remove(base_path)
        except PermissionError:
            ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
            cs1_path = os.path.join(output_dir, f"Step1_Results_{ts}.xlsx")
            print(f"[Step1][WARN] {base_path} 파일을 덮어쓸 수 없어 새 파일로 저장합니다: {cs1_path}")

    with pd.ExcelWriter(cs1_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Original_Data", index=False)
        df_scaled_unweighted.to_excel(writer, sheet_name="Clustered_StandardScaler", index=False)
        df_scaled_weighted.to_excel(writer, sheet_name="Clustered_Weighted", index=False)
        df_counts.to_excel(writer, sheet_name="Cluster_Counts", index=False)
        
        for c in sorted(df_out["cluster"].dropna().unique()):
            cluster_members = df_out[df_out["cluster"] == c][["Lot Number"] + feature_cols]
            cluster_members.to_excel(writer, sheet_name=f"Cluster{int(c)}", index=False)
        
        df_cluster_std.to_excel(writer, sheet_name="Cluster_STD", index=False)
        df_rank.to_excel(writer, sheet_name="Cluster_STD_Rank", index=False)
        df_weights.to_excel(writer, sheet_name="Applied_Weights", index=False)  # 새 시트
        df_compare_summary.to_excel(writer, sheet_name="Cluster_Compare_Summary", index=False)
        df_compare_matrix.to_excel(writer, sheet_name="Cluster_Compare_Contingency")

        x_min, x_max = scatter_plot_df["PC1"].min(), scatter_plot_df["PC1"].max()
        y_min, y_max = scatter_plot_df["PC2"].min(), scatter_plot_df["PC2"].max()

        fig, ax = plt.subplots(figsize=(8, 6))
        clusters_sorted = sorted(scatter_plot_df["cluster"].unique())
        cmap = plt.cm.get_cmap("tab10", len(clusters_sorted))
        cluster_colors = {}
        for idx, cluster_id in enumerate(clusters_sorted):
            cluster_points = scatter_plot_df[scatter_plot_df["cluster"] == cluster_id]
            color = cmap(idx)
            cluster_colors[cluster_id] = color
            ax.scatter(
                cluster_points["PC1"],
                cluster_points["PC2"],
                color=color,
                label=f"Cluster {cluster_id}",
                s=25,
                alpha=0.7,
            )
        ax.set_title("Cluster Distribution (PCA 2D)")
        ax.set_xlabel("PC1")
        ax.set_ylabel("PC2")
        ax.set_xlim(x_min, x_max)
        ax.set_ylim(y_min, y_max)
        ax.legend(loc="best", fontsize=8)
        plt.tight_layout()

        scatter_img = BytesIO()
        plt.savefig(scatter_img, format="png", dpi=150)
        plt.close(fig)
        scatter_img.seek(0)
        scatter_sheet = writer.book.create_sheet("Cluster_Scatter")
        scatter_image = XLImage(scatter_img)
        scatter_image.width, scatter_image.height = 720, 480
        scatter_sheet.add_image(scatter_image, "A1")

        n_clusters = len(clusters_sorted)
        if n_clusters > 0:
            ncols = min(3, n_clusters)
            nrows = int(np.ceil(n_clusters / ncols))
            fig_multi, axes_multi = plt.subplots(nrows, ncols, figsize=(4.5 * ncols, 3.5 * nrows))
            axes_multi = np.array(axes_multi).reshape(-1)
            for ax in axes_multi[n_clusters:]:
                ax.axis("off")
            for ax, cluster_id in zip(axes_multi, clusters_sorted):
                cluster_points = scatter_plot_df[scatter_plot_df["cluster"] == cluster_id]
                ax.scatter(
                    cluster_points["PC1"],
                    cluster_points["PC2"],
                    color=cluster_colors.get(cluster_id, "#1f77b4"),
                    s=30,
                    alpha=0.8,
                )
                ax.set_title(f"Cluster {cluster_id}")
                ax.set_xlabel("PC1")
                ax.set_ylabel("PC2")
                ax.set_xlim(x_min, x_max)
                ax.set_ylim(y_min, y_max)
                ax.grid(True, linestyle="--", alpha=0.3)
            plt.tight_layout()

            cluster_img = BytesIO()
            plt.savefig(cluster_img, format="png", dpi=150)
            plt.close(fig_multi)
            cluster_img.seek(0)

            multi_sheet = writer.book.create_sheet("Cluster_Scatter_ByCluster")
            multi_image = XLImage(cluster_img)
            multi_image.width, multi_image.height = 720, 520
            multi_sheet.add_image(multi_image, "A1")

        _add_k_selection_sheet(writer.book, k_results)
    
    print(f"[Step1] 저장 완료: {cs1_path}")
    return cs1_path, best_cluster, worst_cluster


__all__ = ["run_step1"]
